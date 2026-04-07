#!/usr/bin/env python3
"""
Exam Timetable Generator
Processes an XLSX exam schedule and generates a personalized timetable.
"""

import re
import sys
import csv
import pandas as pd
import openpyxl
from datetime import datetime
from collections import defaultdict

# ─────────────────────────────────────────────
# SECTION 1: NORMALIZATION & PARSING UTILITIES
# ─────────────────────────────────────────────

def normalize_text(text):
    """Strip whitespace, lowercase, remove invisible/noise characters."""
    if text is None:
        return ""
    return re.sub(r'\s+', ' ', str(text)).strip().lower()

def extract_course_codes(text):
    """
    Extract all course codes like cs1004, ee2001, ai3002 from arbitrary text.
    """
    # We use a copy of text here to not affect casing for name extraction later
    search_text = normalize_text(text)
    matches = re.findall(r'\b([a-z]{2,4}\d{3,5})\b', search_text)
    return set(matches)

def extract_course_name(text, codes):
    """
    Extract course name found between the last course code and the first stream marker.
    Example: "SS1015 Pakistan Studies BS(CS)" -> "Pakistan Studies"
    """
    if not codes:
        return "Unknown"
    
    # Find the end position of the last occurring course code in the raw text
    last_code_end = 0
    for code in codes:
        # Case-insensitive search to find the actual position in raw text
        for m in re.finditer(re.escape(code), text, re.I):
            if m.end() > last_code_end:
                last_code_end = m.end()
    
    if last_code_end == 0:
        return "Unknown"

    remaining = text[last_code_end:]
    
    # Look for the stream marker like "BS(" or "BS (" or any Degree prefix
    # Matches: BS(CS), BS (CS), BBA(AI), etc.
    m_stream = re.search(r'\b[A-Z]{2,4}\s*\(', remaining)
    if m_stream:
        name = remaining[:m_stream.start()]
    else:
        # Fallback: stop at first parenthesis or newline
        m_stop = re.search(r'[\(\n]', remaining)
        if m_stop:
            name = remaining[:m_stop.start()]
        else:
            name = remaining

    # Clean up symbols often left between code and name
    name = name.strip().strip(",").strip("-").strip(":").strip()
    return name if name else "Unknown"

def extract_batches(text):
    """Extract all 4-digit year batches like 2023, 2022."""
    text = normalize_text(text)
    matches = re.findall(r'\b(20\d{2})\b', text)
    return set(matches)

def extract_streams(text):
    """
    Extract all stream identifiers (cs, ai, ee, etc.).
    Must be 2-4 uppercase/lowercase letters, not a known noise word or 'bs/bsc'.
    """
    text = normalize_text(text)
    # Common noise words and common degree prefixes like 'bs' or 'bsc'
    noise = {'and', 'the', 'for', 'in', 'of', 'to', 'at', 'by', 'pm', 'am', 'bs', 'bsc'}
    matches = re.findall(r'\b([a-z]{2,4})\b', text)
    streams = set()
    for m in matches:
        if m not in noise and not m.isdigit():
            streams.add(m)
    return streams

def parse_time_slot(text):
    """
    Parse time slots like "9:00–11:00" and return "HH:MM AM/PM - HH:MM AM/PM".
    """
    if not text:
        return "Unknown"
    text = str(text).strip()

    # Pattern: HH:MM [AM/PM] [–/-] HH:MM [AM/PM]
    pattern = r'(\d{1,2}:\d{2})\s*(am|pm)?\s*[-–—to]+\s*(\d{1,2}:\d{2})\s*(am|pm)?'
    m = re.search(pattern, text, re.IGNORECASE)
    if m:
        start, s_ampm, end, e_ampm = m.groups()
        if not s_ampm and e_ampm: s_ampm = e_ampm
        return f"{_format_12h(start, s_ampm)} - {_format_12h(end, e_ampm)}"

    # Pattern: compact like "0900-1100"
    m = re.search(r'\b(\d{4})\s*[-–—]\s*(\d{4})\b', text)
    if m:
        s, e = m.group(1), m.group(2)
        return f"{_format_12h(f'{s[:2]}:{s[2:]}')} - {_format_12h(f'{e[:2]}:{e[2:]}')}"

    # Pattern: integer range like "9-11"
    m = re.search(r'\b(\d{1,2})\s*[-–—]\s*(\d{1,2})\b', text)
    if m:
        h1, h2 = m.group(1), m.group(2)
        return f"{_format_12h(f'{int(h1):02d}:00')} - {_format_12h(f'{int(h2):02d}:00')}"

    return text.strip() or "Unknown"

def _format_12h(time_str, ampm=None):
    """Convert HH:MM [am/pm] to 12-hour format with AM/PM."""
    try:
        h, m = map(int, time_str.split(':'))
        if ampm:
            ampm = ampm.lower()
            if ampm == 'pm' and h != 12: h += 12
            elif ampm == 'am' and h == 12: h = 0
        elif h < 8: # Heuristic: if AM/PM missing and h < 8, it's likely PM
            h += 12
        
        suffix = "AM" if h < 12 else "PM"
        h12 = h % 12
        if h12 == 0: h12 = 12
        return f"{h12:02d}:{m:02d} {suffix}"
    except:
        return time_str

def _time_sort_key(time_str):
    """Extract first 24h-equivalent time for sorting."""
    m = re.search(r'(\d{1,2}):(\d{2})\s*(AM|PM)', time_str, re.I)
    if m:
        h, mm, ampm = int(m.group(1)), int(m.group(2)), m.group(3).upper()
        if ampm == 'PM' and h != 12: h += 12
        elif ampm == 'AM' and h == 12: h = 0
        return f"{h:02d}:{mm:02d}"
    return "23:59"

def parse_date_day(val):
    """
    Parse date and day, handling datetime objects, ISO strings, and messy text.
    Ensures Date is dd/mm/yyyy and Day is the full day name.
    """
    if val is None:
        return "Unknown", "Unknown"
    
    found_date = None
    day_str = "Unknown"

    # Handle actual datetime objects from openpyxl
    if isinstance(val, datetime):
        found_date = val
    else:
        text = str(val).strip()
        # Look for explicit day name first
        day_names = ['monday','tuesday','wednesday','thursday','friday','saturday','sunday',
                     'mon','tue','wed','thu','fri','sat','sun']
        for d in day_names:
            if re.search(r'\b' + d + r'\b', text, re.IGNORECASE):
                full_days = {'mon':'Monday','tue':'Tuesday','wed':'Wednesday','thu':'Thursday',
                             'fri':'Friday','sat':'Saturday','sun':'Sunday'}
                d_lower = d.lower()
                day_str = full_days.get(d_lower[:3], d_lower.capitalize())
                break

        # Try various date patterns
        # ISO YYYY-MM-DD
        m = re.search(r'(\d{4})-(\d{2})-(\d{2})', text)
        if m:
            try: found_date = datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
            except: pass
        
        # DD/MM/YYYY
        if not found_date:
            m = re.search(r'(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})', text)
            if m:
                try: found_date = datetime(int(m.group(3)), int(m.group(2)), int(m.group(1)))
                except: pass

        # Month Name formats
        month_map = {
            'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,
            'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12,
            'january':1,'february':2,'march':3,'april':4,'may':5,'june':6,
            'july':7,'august':8,'september':9,'october':10,'november':11,'december':12
        }
        if not found_date:
            m = re.search(r'(\d{1,2})\s+([a-zA-Z]{3,9})(?:\s+(\d{4}))?', text)
            if m and m.group(2).lower() in month_map:
                d_n, m_s, y = m.group(1), m.group(2).lower(), m.group(3) or str(datetime.now().year)
                try: found_date = datetime(int(y), month_map[m_s], int(d_n))
                except: pass

    if found_date:
        date_str = found_date.strftime("%d/%m/%Y")
        if day_str == "Unknown":
            day_str = found_date.strftime("%A")
        return date_str, day_str
    
    return str(val), "Unknown"

def date_sort_key(date_str):
    """Return a sortable key from DD/MM/YYYY string."""
    m = re.match(r'(\d{2})/(\d{2})/(\d{4})', date_str)
    if m:
        return (int(m.group(3)), int(m.group(2)), int(m.group(1)))
    return (9999, 99, 99)


# ─────────────────────────────────────────────
# SECTION 2: XLSX LOADING & PREPROCESSING
# ─────────────────────────────────────────────

VALID_SHEETS = ['FSC', 'FSM', 'EE']

def load_workbook_sheet(filepath, sheet_name):
    """Load raw cell values from an openpyxl sheet (preserves merged cells)."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
    ws = wb[sheet_name]
    return ws

def preprocess_sheet(ws):
    """
    Build lookup index, handling merged cells for time (row 4) and date (col A).
    """
    time_row = {}
    date_col = {}
    cells = []

    max_row = ws.max_row
    max_col = ws.max_column

    # Pre-map merged cells for fast value retrieval
    merged_map = {}
    for merged_range in ws.merged_cells.ranges:
        top_left_val = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
        for r in range(merged_range.min_row, merged_range.max_row + 1):
            for c in range(merged_range.min_col, merged_range.max_col + 1):
                merged_map[(r, c)] = top_left_val

    def get_val(r, c):
        if (r, c) in merged_map:
            return merged_map[(r, c)]
        return ws.cell(row=r, column=c).value

    # Extract time slots from row 4 (Column B onwards)
    for col in range(2, max_col + 1):
        val = get_val(4, col)
        if val:
            time_row[col] = parse_time_slot(str(val))

    # Extract dates from column A (Row 5 onwards)
    for row in range(5, max_row + 1):
        val = get_val(row, 1)
        if val:
            date_str, day_str = parse_date_day(val)
            date_col[row] = (date_str, day_str)

    # Preprocess data cells
    for row in range(5, max_row + 1):
        for col in range(2, max_col + 1):
            raw = get_val(row, col)
            if raw is None:
                continue
            raw_str = str(raw).strip()
            if not raw_str:
                continue
            
            codes   = extract_course_codes(raw_str)
            batches = extract_batches(raw_str)
            streams = extract_streams(raw_str)
            
            if not codes:
                continue

            # Extract course name from raw string
            name = extract_course_name(raw_str, codes)

            cells.append({
                'row': row,
                'col': col,
                'raw': raw_str,
                'course_codes': codes,
                'course_name': name,
                'batches': batches,
                'streams': streams,
            })

    return time_row, date_col, cells


# ─────────────────────────────────────────────
# SECTION 3: USER INPUT PARSING
# ─────────────────────────────────────────────

def parse_course_input(raw_input, default_batch, default_stream):
    """
    Parse course input in formats:
      course_code
      course_code, batch
      course_code, batch, stream
    Returns (course_code, batch, stream) or None on invalid.
    """
    parts = [p.strip() for p in raw_input.split(',')]
    if not parts:
        return None

    course_code = normalize_text(parts[0])
    # Basic validation for the search input
    if not re.match(r'^[a-z]{2,4}\d{3,5}$', course_code):
        print(f"  ⚠  Invalid course code format: '{parts[0]}'")
        return None

    batch = normalize_text(parts[1]) if len(parts) > 1 and parts[1] else default_batch
    stream = normalize_text(parts[2]) if len(parts) > 2 and parts[2] else default_stream

    # Validate batch
    if not re.match(r'^20\d{2}$', str(batch)):
        print(f"  ⚠  Invalid batch '{batch}', using default '{default_batch}'")
        batch = default_batch

    return course_code, str(batch), str(stream)


# ─────────────────────────────────────────────
# SECTION 4: MATCHING ENGINE
# ─────────────────────────────────────────────

def find_matches(cells, time_row, date_col, course_code, batch, stream):
    """
    Scan preprocessed cells and return matching exam entries.
    Matching priority:
      1. course_code must match (required)
      2. batch match preferred (mandatory if batches are specified in cell)
      3. stream match preferred (mandatory if streams are specified in cell)
    """
    results = []
    seen = set()  # Deduplicate by (row, col)

    for cell in cells:
        # Check if requested course is in this cell's course list
        if course_code not in cell['course_codes']:
            continue

        # Batch matching: if cell has batches listed, the searched batch MUST be one of them
        if cell['batches'] and batch not in cell['batches']:
            continue

        # Stream matching: if cell has streams listed, the searched stream MUST be one of them
        if cell['streams'] and stream not in cell['streams']:
            continue

        key = (cell['row'], cell['col'])
        if key in seen:
            continue
        seen.add(key)

        row, col = cell['row'], cell['col']
        time_slot = time_row.get(col, "Unknown")
        date_str, day_str = date_col.get(row, ("Unknown", "Unknown"))

        results.append({
            'date': date_str,
            'day': day_str,
            'time': time_slot,
            'course_code': course_code.upper() if isinstance(course_code, str) else ", ".join(sorted(list(cell['course_codes']))).upper(),
            'course_name': cell['course_name'],
            'batch': batch,
            'stream': stream.upper(),
            'raw': cell['raw'],
        })

    return results


def find_all_for_batch_stream(cells, time_row, date_col, batch, stream):
    """
    Automatically find all exam entries for a specific batch and stream.
    """
    results = []
    seen = set()
    for cell in cells:
        # Match if both batch and stream are explicitly mentioned in the cell
        if batch in cell['batches'] and stream in cell['streams']:
            key = (cell['row'], cell['col'])
            if key in seen:
                continue
            seen.add(key)

            row, col = cell['row'], cell['col']
            time_slot = time_row.get(col, "Unknown")
            date_str, day_str = date_col.get(row, ("Unknown", "Unknown"))

            # Extract all course codes found in this cell
            codes = sorted(list(cell['course_codes']))
            code_str = ", ".join(codes).upper() if codes else "Unknown"

            results.append({
                'date': date_str,
                'day': day_str,
                'time': time_slot,
                'course_code': code_str,
                'course_name': cell['course_name'],
                'batch': batch,
                'stream': stream.upper(),
                'raw': cell['raw'],
            })
    return results

def prompt_mode():
    """Ask user to choose between Default or Custom courses."""
    print("\nSelect Search Mode:")
    print("  1. Default Courses (Automatic: all exams for your Batch & Stream)")
    print("  2. Custom Courses  (Manual: enter specific course codes)")
    while True:
        choice = input("\nChoice (1 or 2): ").strip()
        if choice == '1':
            return 'DEFAULT'
        if choice == '2':
            return 'CUSTOM'
        print("  ⚠  Invalid choice. Please enter 1 or 2.")

# ─────────────────────────────────────────────
# SECTION 5: OUTPUT FORMATTING & EXPORT
# ─────────────────────────────────────────────

def display_timetable(entries):
    """Print a formatted timetable to stdout."""
    if not entries:
        print("\n  No exam entries to display.")
        return

    # Sort by date then time
    sorted_entries = sorted(entries, key=lambda e: (date_sort_key(e['date']), _time_sort_key(e['time'])))

    col_w = [10, 10, 18, 10, 25, 8, 8]
    headers = ['Date', 'Day', 'Time', 'Code', 'Course Name', 'Batch', 'Stream']
    sep = '+' + '+'.join('-' * (w + 2) for w in col_w) + '+'
    row_fmt = '| ' + ' | '.join(f'{{:<{w}}}' for w in col_w) + ' |'

    print("\n" + "═" * len(sep))
    print("  📅  PERSONALIZED EXAM TIMETABLE")
    print("═" * len(sep))
    print(sep)
    print(row_fmt.format(*headers))
    print(sep)
    for e in sorted_entries:
        print(row_fmt.format(
            e['date'][:col_w[0]],
            e['day'][:col_w[1]],
            e['time'][:col_w[2]],
            e['course_code'][:col_w[3]],
            e['course_name'][:col_w[4]],
            str(e['batch'])[:col_w[5]],
            e['stream'][:col_w[6]],
        ))
    print(sep)
    print(f"\n  Total exams found: {len(sorted_entries)}")

def export_csv(entries, filepath='output.csv'):
    """Export timetable to CSV."""
    if not entries:
        return
    sorted_entries = sorted(entries, key=lambda e: (date_sort_key(e['date']), _time_sort_key(e['time'])))
    fieldnames = ['Date', 'Day', 'Time', 'Course Code', 'Course Name', 'Batch', 'Stream']
    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for e in sorted_entries:
            writer.writerow({
                'Date': e['date'], 'Day': e['day'], 'Time': e['time'],
                'Course Code': e['course_code'], 'Course Name': e['course_name'],
                'Batch': e['batch'], 'Stream': e['stream'],
            })
    print(f"  ✅  CSV exported → {filepath}")

def export_tsv(entries, filepath='output.tsv'):
    """Export timetable to TSV."""
    if not entries:
        return
    sorted_entries = sorted(entries, key=lambda e: (date_sort_key(e['date']), _time_sort_key(e['time'])))
    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f, delimiter='\t')
        writer.writerow(['Date', 'Day', 'Time', 'Course Code', 'Course Name', 'Batch', 'Stream'])
        for e in sorted_entries:
            writer.writerow([e['date'], e['day'], e['time'],
                             e['course_code'], e['course_name'], e['batch'], e['stream']])
    print(f"  ✅  TSV exported → {filepath}")

def export_xlsx(entries, filepath='output.xlsx'):
    """Bonus: Export timetable to XLSX with basic formatting."""
    if not entries:
        return
    from openpyxl.styles import Font, PatternFill, Alignment

    sorted_entries = sorted(entries, key=lambda e: (date_sort_key(e['date']), _time_sort_key(e['time'])))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Exam Timetable"

    headers = ['Date', 'Day', 'Time', 'Course Code', 'Course Name', 'Batch', 'Stream']
    header_fill = PatternFill('solid', start_color='2E4057')
    header_font = Font(bold=True, color='FFFFFF', name='Arial', size=11)
    data_font = Font(name='Arial', size=10)
    alt_fill = PatternFill('solid', start_color='EBF5FB')

    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for i, e in enumerate(sorted_entries):
        row = [e['date'], e['day'], e['time'], e['course_code'], e['course_name'], e['batch'], e['stream']]
        ws.append(row)
        for cell in ws[i + 2]:
            cell.font = data_font
            cell.alignment = Alignment(horizontal='center')
            if i % 2 == 1:
                cell.fill = alt_fill

    col_widths = [14, 12, 22, 14, 30, 10, 10]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    wb.save(filepath)
    print(f"  ✅  XLSX exported → {filepath}")


# ─────────────────────────────────────────────
# SECTION 6: MAIN CLI FLOW
# ─────────────────────────────────────────────

def prompt_school():
    """Prompt user to select a school/sheet."""
    print("\n┌─────────────────────────────────┐")
    print("│   EXAM TIMETABLE GENERATOR      │")
    print("└─────────────────────────────────┘")
    print("\nAvailable schools/sheets:", ', '.join(VALID_SHEETS))
    while True:
        school = input("Select school (FSC / FSM / EE): ").strip().upper()
        if school in VALID_SHEETS:
            return school
        print(f"  ⚠  Invalid choice. Please enter one of: {', '.join(VALID_SHEETS)}")

def prompt_defaults():
    """Prompt for default batch and stream."""
    while True:
        batch = input("Enter your default batch (e.g., 2023): ").strip()
        if re.match(r'^20\d{2}$', batch):
            break
        print("  ⚠  Batch must be a 4-digit year like 2023.")

    stream = input("Enter your default stream (e.g., CS, AI, EE): ").strip().lower()
    if not stream:
        stream = "unknown"
    return batch, stream

def prompt_filepath():
    """Prompt for XLSX file path."""
    while True:
        path = input("Enter path to the XLSX exam schedule: ").strip()
        try:
            open(path, 'rb').close()
            return path
        except FileNotFoundError:
            print(f"  ⚠  File not found: '{path}'. Please try again.")
        except Exception as e:
            print(f"  ⚠  Cannot open file: {e}")

def main():
    print("\n" + "=" * 50)
    print("  Exam Timetable Generator")
    print("=" * 50)

    # Step 1: Get file path
    filepath = prompt_filepath()

    # Step 2: Select school
    school = prompt_school()

    # Step 3: Get defaults
    default_batch, default_stream = prompt_defaults()

    print(f"\n  Loading sheet '{school}' from '{filepath}'...")
    try:
        ws = load_workbook_sheet(filepath, school)
    except ValueError as e:
        print(f"\n  ❌  Error: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"\n  ❌  Failed to load workbook: {e}")
        sys.exit(1)

    print("  Preprocessing sheet (building lookup index)...")
    time_row, date_col, cells = preprocess_sheet(ws)
    print("  ✅  Indexed {len(cells)} data cells, "
          f"{len(time_row)} time slots, {len(date_col)} date rows.\n")

    # Step 4: Mode selection
    mode = prompt_mode()

    all_entries = []

    if mode == 'DEFAULT':
        print(f"\n  🔍  Scanning for all exams: batch={default_batch}, stream={default_stream.upper()}...")
        all_entries = find_all_for_batch_stream(cells, time_row, date_col, default_batch, default_stream)
        print(f"  ✅  Found {len(all_entries)} regular exam(s).")
    else:
        print("─" * 50)
        print("  Enter courses below. Type -1 when done.")
        print("  Formats: course_code  |  code,batch  |  code,batch,stream")
        print("─" * 50)

        while True:
            raw = input("\nCourse input: ").strip()

            if raw == '-1':
                break

            if not raw:
                print("  ⚠  Empty input. Please enter a course code or -1 to finish.")
                continue

            parsed = parse_course_input(raw, default_batch, default_stream)
            if parsed is None:
                continue

            course_code, batch, stream = parsed
            print(f"  🔍  Searching: course={course_code.upper()}, batch={batch}, stream={stream.upper()}")

            matches = find_matches(cells, time_row, date_col, course_code, batch, stream)

            if not matches:
                print(f"  ⚠  No matches found for '{course_code.upper()}' "
                      f"(batch={batch}, stream={stream.upper()})")
            else:
                print(f"  ✅  Found {len(matches)} exam slot(s):")
                for m in matches:
                    print(f"      {m['date']}  {m['day']}  {m['time']}")
                all_entries.extend(matches)

    # Deduplicate all_entries by (date, time, course_code, batch, stream)
    seen_keys = set()
    deduped = []
    for e in all_entries:
        key = (e['date'], e['time'], e['course_code'], e['batch'], e['stream'])
        if key not in seen_keys:
            seen_keys.add(key)
            deduped.append(e)

    # Display and export
    display_timetable(deduped)

    if deduped:
        print("\n  Exporting...")
        export_csv(deduped, 'output.csv')
        export_tsv(deduped, 'output.tsv')
        export_xlsx(deduped, 'output.xlsx')
    else:
        print("\n  Nothing to export.")

    print("\n  Done. Goodbye! 👋\n")

if __name__ == '__main__':
    main()
