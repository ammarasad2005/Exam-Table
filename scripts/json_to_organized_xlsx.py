#!/usr/bin/env python3
"""
Generate the ORGANIZED xlsx file (exam_schedule_summer.xlsx format)
from the structured summer_schedule.json.

This is the inverse of transform_summer_exam.py — it takes the canonical
JSON and produces the flat-table xlsx that parse-summer-exam.ts expects.

Output format (mirrors the existing repo-root exam_schedule_summer.xlsx):
  - 3 sheets: 'FSC Final', 'FSM Final', 'FSE Final'
  - 8 columns: S.No | Date | Time Slot | Course Code | Course Name | Degree & Sections | Batch | Room
  - Row 1: Title 'Structured Exam Schedule - <SCHOOL> Final Summer 2026' (merged across A1:H1)
  - Row 2: empty
  - Row 3: Headers
  - Row 4+: Data
  - Date format: YY-MM-DD (e.g., '26-08-10' for 2026-08-10)
  - Time format: matches JSON verbatim
  - Degree & Sections: 'ALL' if sections='', else 'ALL (<sections>)'

Usage:
    python3 scripts/json_to_organized_xlsx.py
    python3 scripts/json_to_organized_xlsx.py --input <json> --output <xlsx>
"""
from __future__ import annotations

import argparse
import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

DEFAULT_INPUT = "public/data/summer_schedule.json"
DEFAULT_OUTPUT = "exam_schedule_summer.xlsx"

HEADERS = ["S.No", "Date", "Time Slot", "Course Code", "Course Name",
           "Degree & Sections", "Batch", "Room"]

# Sheet order (matches existing file)
SHEET_ORDER = ["FSC", "FSM", "FSE"]


def format_date_yy_mm_dd(date_str: str) -> str:
    """'10/08/2026' → '26-08-10' (YY-MM-DD, matching existing organized xlsx format)."""
    dd, mm, yyyy = date_str.split("/")
    yy = yyyy[2:]  # last 2 digits of year
    return f"{yy}-{mm}-{dd}"


def format_degree_sections(department: str, sections: str) -> str:
    """Build 'Degree & Sections' cell value.

    'ALL' + ''        → 'ALL'
    'ALL' + 'A'       → 'ALL (A)'
    'ALL' + 'BAF-9A, 9B' → 'ALL (BAF-9A, 9B)'
    """
    if not sections:
        return department
    return f"{department} ({sections})"


def build_sheet(ws, school: str, entries: list[dict]) -> None:
    """Populate a single sheet with the school's entries."""
    # Styling
    title_font = Font(bold=True, size=12)
    header_font = Font(bold=True, size=11)
    body_font = Font(size=11)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # Row 1: Title (merged A1:H1)
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = f"Structured Exam Schedule - {school} Final Summer 2026"
    title_cell.font = title_font
    title_cell.alignment = center_align

    # Row 2: empty
    # Row 3: Headers
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = center_align
        cell.fill = header_fill
        cell.border = thin_border

    # Row 4+: Data
    for i, entry in enumerate(entries, start=1):
        row_idx = 3 + i
        values = [
            i,                                          # S.No (as int)
            format_date_yy_mm_dd(entry["date"]),        # Date
            entry["time"],                              # Time Slot
            entry["courseCode"],                         # Course Code
            entry["courseName"],                         # Course Name
            format_degree_sections(entry["department"], entry["sections"]),  # Degree & Sections
            entry["batch"],                             # Batch
            entry["room"],                              # Room
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = body_font
            cell.border = thin_border
            # Center-align short columns; left-align longer text columns
            if col_idx in (1, 2, 3, 4, 7):  # S.No, Date, Time Slot, Course Code, Batch
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    # Column widths (auto-fit-ish)
    col_widths = {
        1: 6,    # S.No
        2: 12,   # Date
        3: 22,   # Time Slot
        4: 12,   # Course Code
        5: 40,   # Course Name
        6: 22,   # Degree & Sections
        7: 9,    # Batch
        8: 30,   # Room
    }
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze top 3 rows so the title + headers stay visible when scrolling
    ws.freeze_panes = "A4"


def main(argv=None):
    parser = argparse.ArgumentParser(description="Generate organized xlsx from summer_schedule.json")
    parser.add_argument("--input", default=DEFAULT_INPUT, help=f"Input JSON path (default: {DEFAULT_INPUT})")
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help=f"Output xlsx path (default: {DEFAULT_OUTPUT})")
    args = parser.parse_args(argv)

    input_path = Path(args.input)
    output_path = Path(args.output)

    if not input_path.exists():
        print(f"❌ Input file not found: {input_path}")
        return 1

    print(f"📖 Loading JSON: {input_path}")
    entries = json.loads(input_path.read_text(encoding="utf-8"))
    print(f"   Loaded {len(entries)} entries")

    # Group by school
    by_school: dict[str, list[dict]] = {s: [] for s in SHEET_ORDER}
    for e in entries:
        school = e["school"]
        if school not in by_school:
            by_school[school] = []
        by_school[school].append(e)

    print()
    for school in SHEET_ORDER:
        print(f"   {school}: {len(by_school[school])} entries")

    # Build workbook
    print(f"\n🔨 Building workbook...")
    wb = Workbook()
    # Remove the default sheet created by openpyxl
    default_sheet = wb.active
    wb.remove(default_sheet)

    for school in SHEET_ORDER:
        sheet_name = f"{school} Final"
        ws = wb.create_sheet(title=sheet_name)
        build_sheet(ws, school, by_school[school])
        print(f"   ✓ Sheet '{sheet_name}' built with {len(by_school[school])} data rows")

    # Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"\n✅ Wrote organized xlsx to: {output_path}")
    print(f"   File size: {output_path.stat().st_size:,} bytes")
    print(f"   Sheets: {wb.sheetnames}")
    return 0


if __name__ == "__main__":
    import sys
    sys.exit(main())
