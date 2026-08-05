#!/usr/bin/env python3
"""
Phase 4 — Summer Exam Schedule Transformation Script

Reads the RAW messy Excel file (2D grid layout) and writes the structured
`summer_schedule.json` consumed by the FAST-NUCES ISB exam-table frontend.

Input  : /home/z/my-project/upload/Final Examination Schedule Summer 2026 Ver-Draft as on 04-08-2026.xlsx
Output : /home/z/my-project/repo/exam-table/public/data/summer_schedule.json

Re-runnable (idempotent): re-running on the same input produces the same output.
Safety guard: never overwrites the existing JSON with an empty array.

Mapping specification: /home/z/my-project/download/artifact_3_mapping_specification.md
User decisions: /home/z/my-project/download/artifact_3_addendum_user_decisions.md

Usage:
    python3 /home/z/my-project/scripts/transform_summer_exam.py
    python3 /home/z/my-project/scripts/transform_summer_exam.py --input <path> --output <path>
    python3 /home/z/my-project/scripts/transform_summer_exam.py --dry-run   # validate only, don't write
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import sys
from copy import copy
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# ─── Constants (from Artifact 3 addendum) ────────────────────────────────────

# Paths are repo-relative so the script works after `git clone` from any directory.
# Run from the repo root: `python3 scripts/transform_summer_exam.py`
DEFAULT_INPUT = "exam_schedule_summer_raw.xlsx"
DEFAULT_OUTPUT = "public/data/summer_schedule.json"

# Time slot strings (per user Decision #1: explicit AM/PM both ends)
MORNING_TIME = "9:00 AM to 12:00 PM"
AFTERNOON_TIME = "1:00 PM to 4:00 PM"

# Sheet config: sheet_name_prefix → (school_code, has_afternoon_col)
SHEET_CONFIG = [
    ("FSC", "FSC", True),   # FSC (Final) → FSC, has afternoon
    ("FSM", "FSM", True),   # FSM (Final) → FSM, has afternoon
    ("FSE", "FSE", False),  # FSE (Final) → FSE, no afternoon
]

# Column indices (1-based)
DATE_COL = 1      # A
VENUE_COL = 2     # B
MORNING_COL = 3   # C
AFTERNOON_COL = 5 # E

# Validation regexes
RE_DATE = re.compile(r"^\d{2}/\d{2}/\d{4}$")
RE_TIME = re.compile(r"^\d{1,2}:\d{2}\s+(AM|PM)\s+to\s+\d{1,2}:\d{2}\s+(AM|PM)$")
RE_COURSE_CODE = re.compile(r"^[A-Z]{2,4}\d{4}$")
RE_FSM_SECTION = re.compile(r"^([A-Z]{2,5}-\d{1,2}[A-Z]?)$")  # BAF-9A, BFT-09A, MSBA-9A, BBA-9A
RE_FSC_FSE_SECTION = re.compile(r"^([A-Z]{1,2})$")  # A, AB
DAYS_SET = {"Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"}
SCHOOLS_SET = {"FSC", "FSM", "FSE"}


# ─── Excel helpers ────────────────────────────────────────────────────────────

def build_effective_value_map(ws) -> dict[tuple[int, int], Any]:
    """Build a {(row, col): value} dict that resolves merged-cell ranges.

    For every cell in a merged range, the dict returns the top-left cell's value.
    For non-merged cells, it returns the cell's own value.

    This avoids the read-only 'MergedCell' issue — we never mutate the worksheet.
    """
    # Start with all cell values
    value_map: dict[tuple[int, int], Any] = {}
    for row in ws.iter_rows():
        for cell in row:
            value_map[(cell.row, cell.column)] = cell.value

    # Overlay merge ranges: fill every cell in each range with the top-left's value
    for merge in ws.merged_cells.ranges:
        min_row, min_col = merge.min_row, merge.min_col
        top_left_value = value_map.get((min_row, min_col))
        for r in range(min_row, merge.max_row + 1):
            for c in range(min_col, merge.max_col + 1):
                value_map[(r, c)] = top_left_value

    return value_map


def find_header_row(ws, max_scan_row: int = 10) -> Optional[int]:
    """Find the header row by looking for 'Days' in column A and 'Venue' in column B.

    FSC/FSM have headers on row 4; FSE has them on row 5 (row 4 has a whitespace spacer).
    """
    for r in range(1, max_scan_row + 1):
        a_val = ws.cell(row=r, column=DATE_COL).value
        b_val = ws.cell(row=r, column=VENUE_COL).value
        if (a_val and isinstance(a_val, str) and "days" in a_val.lower()
                and b_val and isinstance(b_val, str) and "venue" in b_val.lower()):
            return r
    return None


def find_merge_range_for_cell(ws, row: int, col: int) -> tuple[int, int, int, int]:
    """Return (min_row, min_col, max_row, max_col) of the merge containing (row, col),
    or just (row, col, row, col) if not in any merge.
    """
    for merge in ws.merged_cells.ranges:
        if (merge.min_row <= row <= merge.max_row
                and merge.min_col <= col <= merge.max_col):
            return merge.min_row, merge.min_col, merge.max_row, merge.max_col
    return row, col, row, col


def collect_rooms_for_merge(value_map, top_row: int, bottom_row: int, col: int = VENUE_COL) -> list[str]:
    """Collect room values from column B (or specified col) for rows top_row..bottom_row.

    Uses the pre-built value_map (which already resolves merged cells).
    Skips empty room values. Returns list of room strings in source order (top to bottom).
    """
    rooms = []
    for r in range(top_row, bottom_row + 1):
        v = value_map.get((r, col))
        if v is not None and str(v).strip():
            rooms.append(str(v).strip())
    return rooms


def process_sheet(ws, school: str, has_afternoon: bool) -> list[dict]:
    """Walk a single sheet's grid and emit ExamEntry dicts.

    Returns a list of entries (unsorted).
    """
    # Build effective value map (resolves merged cells without mutating the worksheet)
    value_map = build_effective_value_map(ws)

    header_row = find_header_row(ws)
    if header_row is None:
        print(f"  ⚠️  No header row found in sheet '{ws.title}' — skipping.", file=sys.stderr)
        return []

    entries: list[dict] = []
    current_date_str: Optional[str] = None
    current_day: Optional[str] = None

    # Iterate rows from header+1 onwards
    for r in range(header_row + 1, ws.max_row + 1):
        cell_a = value_map.get((r, DATE_COL))

        # Detect date group start (cell A is a datetime — could be top-left of merge
        # OR inherited from a merge). Either way, update current_date_str.
        if isinstance(cell_a, dt.datetime):
            current_date_str = cell_a.strftime("%d/%m/%Y")
            current_day = cell_a.strftime("%A")
            # IMPORTANT: do NOT `continue` — the top-left row of a date merge ALSO has
            # course cells (column B/C/E) that need to be processed in this same row.
            # Fall through to course parsing below.

        # Detect end-of-data (cell A is a string starting with "Note" or "\nNote")
        elif isinstance(cell_a, str):
            stripped = cell_a.strip().lower()
            if stripped.startswith("note:") or stripped.startswith("\nnote:") or "students are informed" in stripped:
                break
            # Empty string or whitespace-only → could be a separator row; skip but keep current date
            if not stripped:
                continue
            # Other string in column A? Unlikely — log and skip
            print(f"  ⚠️  Unexpected string in column A row {r}: {cell_a!r}", file=sys.stderr)
            continue

        # else: cell_a is None — could be a separator row or a row inside a date group
        # (None means no merge inheritance either). Fall through to course parsing if we
        # have a current_date_str.

        # If we're inside a date group, process course cells
        if current_date_str is None:
            # No date set yet — skip pre-data rows
            continue

        # Morning course (column C)
        morning_val = value_map.get((r, MORNING_COL))
        if morning_val is not None and str(morning_val).strip():
            morning_text = str(morning_val).strip()
            # Find the merge range for this course cell (to collect rooms)
            mr_min, mc_min, mr_max, mc_max = find_merge_range_for_cell(ws, r, MORNING_COL)
            # Only process if this is the top-left of the merge (or a standalone single cell)
            if mr_min == r and mc_min == MORNING_COL:
                rooms = collect_rooms_for_merge(value_map, mr_min, mr_max)
                entry = build_entry(current_date_str, current_day, MORNING_TIME,
                                    morning_text, rooms, school)
                if entry:
                    entries.append(entry)

        # Afternoon course (column E) — only FSC and FSM
        if has_afternoon:
            afternoon_val = value_map.get((r, AFTERNOON_COL))
            if afternoon_val is not None and str(afternoon_val).strip():
                afternoon_text = str(afternoon_val).strip()
                mr_min, mc_min, mr_max, mc_max = find_merge_range_for_cell(ws, r, AFTERNOON_COL)
                if mr_min == r and mc_min == AFTERNOON_COL:
                    rooms = collect_rooms_for_merge(value_map, mr_min, mr_max)
                    entry = build_entry(current_date_str, current_day, AFTERNOON_TIME,
                                        afternoon_text, rooms, school)
                    if entry:
                        entries.append(entry)

    return entries

def normalize_course_code(raw_token: str) -> str:
    """Strip hyphens and spaces from a course code token. Uppercase.

    'MT-1003' → 'MT1003'
    'EL 1005' → 'EL1005'
    'CS1004'  → 'CS1004'
    """
    return re.sub(r"[\-\s]", "", raw_token).upper()


def split_code_and_rest(cell_text: str) -> tuple[str, str]:
    """Split a course cell into (course_code, rest_after_code).

    The course code is the leading token, which may itself contain a hyphen or
    space between the letter prefix and the 4-digit number:
      'MT-1003 Calculus and Analytical Geometry'
      'EL 1005 Digital Logic Design Lab'
      'CS1004 Object Oriented Programming'
      'AF1002 Financial Accounting \\nBAF-9A, 9B'

    Pattern: ^[A-Z]{2,4}[-\\s]?\\d{4}  (letters, optional hyphen/space, 4 digits)

    Also strips a leading ' - ' separator from the rest if present.

    Returns the NORMALIZED code (hyphens/spaces stripped, uppercased).
    """
    # Match the course code at the start of the string.
    # Pattern: 2-4 uppercase letters, optional hyphen or space, 4 digits.
    m = re.match(r"^\s*([A-Za-z]{2,4}[-\s]?\d{4})\s+(.*)$", cell_text, re.DOTALL)
    if not m:
        # Maybe the whole cell is just a code (no name) — try matching just the code
        m2 = re.match(r"^\s*([A-Za-z]{2,4}[-\s]?\d{4})\s*$", cell_text)
        if m2:
            return normalize_course_code(m2.group(1)), ""
        # Fall back to old behavior (split on first whitespace)
        parts = cell_text.strip().split(None, 1)
        if len(parts) == 2:
            return normalize_course_code(parts[0]), parts[1].strip()
        return normalize_course_code(cell_text.strip()), ""

    code_raw, rest = m.group(1), m.group(2)
    code = normalize_course_code(code_raw)
    rest = rest.strip()
    # Strip leading ' - ' separator if present (e.g., 'EE2008 - Signals and Systems')
    rest = re.sub(r"^[-\u2013\u2014]\s+", "", rest)
    return code, rest


def parse_course_cell_fsc(cell_text: str) -> tuple[str, str]:
    """Parse an FSC course cell. Sections are ALWAYS dropped (user Decision #2).

    Per Decision #2: strip a SINGLE trailing capital letter (preceded by whitespace)
    and set sections="". Do NOT strip 2-letter trailing groups — those are part of
    the course name (e.g., "Generative AI" keeps "AI"; "Discrete Structures A" drops "A").

    'MT-1003 Calculus and Analytical Geometry' → ('MT1003', 'Calculus and Analytical Geometry')
    'CS1002 Programming Fundamental A'           → ('CS1002', 'Programming Fundamental')
    'AI4009 Generative AI'                       → ('AI4009', 'Generative AI')  ← keeps 'AI'
    'EL 1005 Digital Logic Design Lab'           → ('EL1005', 'Digital Logic Design Lab')
    'CL 1002 Programming Fundamental Lab A'      → ('CL1002', 'Programming Fundamental Lab')
    """
    code, rest = split_code_and_rest(cell_text)
    # Strip a SINGLE trailing capital letter preceded by whitespace (e.g., ' A', ' B')
    # Use lookbehind to ensure the letter is NOT part of a longer uppercase sequence.
    # Match: whitespace + single uppercase letter + end-of-string
    rest = re.sub(r"\s+([A-Z])\s*$", "", rest).strip()
    return code, rest


def parse_course_cell_fsm(cell_text: str) -> tuple[str, str, str]:
    """Parse an FSM course cell. Returns (code, name, sections).

    Two patterns:
      1. Newline-separated: 'AF1002 Financial Accounting \nBAF-9A, 9B'
         → name='Financial Accounting', sections='BAF-9A, 9B'
      2. Inline section code: 'CS2016 - Programming for Business BFT-09A'
         → name='Programming for Business', sections='BFT-09A'
    """
    code, rest = split_code_and_rest(cell_text)

    # Pattern 1: newline separator
    if "\n" in rest:
        parts = rest.split("\n", 1)
        name = parts[0].strip()
        sections = parts[1].strip() if len(parts) > 1 else ""
        return code, name, sections

    # Pattern 2: trailing inline section code (e.g., 'BFT-09A', 'MSBA-9A')
    # Match: <name> <section_code> at end of string
    m = re.search(r"^(.*?)\s+([A-Z]{2,5}-\d{1,2}[A-Z]?)\s*$", rest)
    if m:
        return code, m.group(1).strip(), m.group(2)

    # No section found
    return code, rest.strip(), ""


def parse_course_cell_fse(cell_text: str) -> tuple[str, str, str]:
    """Parse an FSE course cell. Returns (code, name, sections).

    'EE2008 - Signals and Systems A'         → ('EE2008', 'Signals and Systems', 'A')
    'EE1001 - Linear Circuit Anlysis.  AB'    → ('EE1001', 'Linear Circuit Anlysis', 'AB')
    'MT1006 Differential Equations AB'        → ('MT1006', 'Differential Equations', 'AB')
    """
    code, rest = split_code_and_rest(cell_text)
    # Strip trailing section letters (1-2 uppercase chars at end, optionally preceded by period+space)
    # Use a non-greedy match for the name to allow the section to be the LAST 1-2 caps
    # Also handle the case where there's a trailing period and double space (e.g., 'Anlysis.  AB')
    m = re.search(r"^(.*?)\s*[.\s]?\s*([A-Z]{1,2})\s*$", rest)
    if m:
        name = m.group(1).strip().rstrip(".").strip()
        section = m.group(2)
        # Sanity: don't extract if 'name' would be empty (e.g., entire rest was just 'A')
        if name:
            return code, name, section
    return code, rest.strip(), ""


# ─── Grid walker (see process_sheet / collect_rooms_for_merge / build_entry above) ───


def build_entry(date_str: str, day: str, time_str: str, course_cell_text: str,
                rooms: list[str], school: str) -> Optional[dict]:
    """Build an ExamEntry dict from components. Returns None if course cell is empty."""
    if not course_cell_text.strip():
        return None

    if school == "FSC":
        code, name = parse_course_cell_fsc(course_cell_text)
        sections = ""
    elif school == "FSM":
        code, name, sections = parse_course_cell_fsm(course_cell_text)
    elif school == "FSE":
        code, name, sections = parse_course_cell_fse(course_cell_text)
    else:
        print(f"  ⚠️  Unknown school: {school}", file=sys.stderr)
        return None

    room_str = ", ".join(rooms)

    return {
        "date": date_str,
        "day": day,
        "time": time_str,
        "courseCode": code,
        "courseName": name,
        "batch": "Summer",
        "department": "ALL",
        "school": school,
        "room": room_str,
        "sections": sections,
    }


# ─── Sorting & Validation ─────────────────────────────────────────────────────

def parse_date_for_sort(date_str: str) -> tuple[int, int, int]:
    """Parse 'DD/MM/YYYY' → (year, month, day) for sorting."""
    d, m, y = date_str.split("/")
    return int(y), int(m), int(d)


def parse_time_for_sort(time_str: str) -> int:
    """Parse '9:00 AM to 12:00 PM' → minutes from midnight (uses FIRST time).

    Mirrors the frontend parseTime() regex: /(\\d{1,2}):(\\d{2})\\s*(AM|PM)/i
    """
    m = re.match(r"^\s*(\d{1,2}):(\d{2})\s*(AM|PM)", time_str, re.IGNORECASE)
    if not m:
        return 0
    h = int(m.group(1))
    mins = int(m.group(2))
    period = m.group(3).upper()
    if period == "PM" and h < 12:
        h += 12
    if period == "AM" and h == 12:
        h = 0
    return h * 60 + mins


def sort_entries(entries: list[dict]) -> list[dict]:
    """Sort by (date asc, time asc)."""
    return sorted(entries, key=lambda e: (parse_date_for_sort(e["date"]), parse_time_for_sort(e["time"])))


def validate_entries(entries: list[dict]) -> list[str]:
    """Run validation checks. Returns list of error/warning messages (empty if all pass).

    Per Artifact 3 §8 + addendum (room+time conflict check REMOVED — shared rooms are deliberate).
    """
    errors: list[str] = []
    warnings: list[str] = []

    if not entries:
        errors.append("No entries parsed — refusing to write empty array (safety guard).")
        return errors

    if len(entries) < 20:
        errors.append(f"Sanity check failed: only {len(entries)} entries (expected ~27).")

    seen_dedupe_keys = set()
    duplicate_keys = []
    for i, e in enumerate(entries):
        prefix = f"Entry #{i} ({e.get('date','?')} {e.get('courseCode','?')})"
        # date
        if not RE_DATE.match(e.get("date", "")):
            errors.append(f"{prefix}: date '{e.get('date')}' does not match DD/MM/YYYY")
        # day
        if e.get("day") not in DAYS_SET:
            errors.append(f"{prefix}: day '{e.get('day')}' not a valid weekday")
        # time
        if not RE_TIME.match(e.get("time", "")):
            errors.append(f"{prefix}: time '{e.get('time')}' does not match 'H:MM AM/PM to H:MM AM/PM'")
        # courseCode
        if not RE_COURSE_CODE.match(e.get("courseCode", "")):
            warnings.append(f"{prefix}: courseCode '{e.get('courseCode')}' does not match ^[A-Z]{{2,4}}\\d{{4}}$")
        # courseName
        if not e.get("courseName"):
            errors.append(f"{prefix}: courseName is empty")
        # batch
        if e.get("batch") != "Summer":
            errors.append(f"{prefix}: batch is '{e.get('batch')}', expected 'Summer'")
        # department
        if e.get("department") != "ALL":
            errors.append(f"{prefix}: department is '{e.get('department')}', expected 'ALL'")
        # school
        if e.get("school") not in SCHOOLS_SET:
            errors.append(f"{prefix}: school '{e.get('school')}' not in FSC/FSM/FSE")
        # room
        if not e.get("room"):
            errors.append(f"{prefix}: room is empty (every exam needs ≥1 room)")
        # dedupe
        dedupe_key = (e["date"], e["time"], e["courseCode"], e["sections"])
        if dedupe_key in seen_dedupe_keys:
            duplicate_keys.append(f"{prefix}: duplicate (date, time, courseCode, sections)")
        seen_dedupe_keys.add(dedupe_key)

    if duplicate_keys:
        for d in duplicate_keys:
            warnings.append(f"Duplicate entry: {d}")

    # Sort order check (entries must already be sorted by date+time)
    sorted_copy = sort_entries(entries)
    if sorted_copy != entries:
        errors.append("Entries are not sorted by (date, time) — auto-sorting before write.")

    return errors + warnings


# ─── Main ──────────────────────────────────────────────────────────────────────

def main(argv=None):
    parser = argparse.ArgumentParser(description="Transform raw summer exam xlsx → summer_schedule.json")
    parser.add_argument("--input", default=DEFAULT_INPUT, help=f"Input xlsx path (default: {DEFAULT_INPUT})")
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help=f"Output JSON path (default: {DEFAULT_OUTPUT})")
    parser.add_argument("--dry-run", action="store_true", help="Validate only — do not write output file")
    args = parser.parse_args(argv)

    input_path = Path(args.input)
    output_path = Path(args.output)

    if not input_path.exists():
        print(f"❌ Input file not found: {input_path}", file=sys.stderr)
        return 1

    print(f"📖 Loading workbook: {input_path}")
    wb = load_workbook(input_path, data_only=True)
    print(f"   Sheets: {wb.sheetnames}")

    all_entries: list[dict] = []
    for sheet_prefix, school_code, has_afternoon in SHEET_CONFIG:
        # Find sheet whose name starts with the prefix (e.g., 'FSC (Final)' matches 'FSC')
        sheet_name = None
        for name in wb.sheetnames:
            if name.upper().startswith(sheet_prefix.upper()):
                sheet_name = name
                break
        if sheet_name is None:
            print(f"  ⚠️  No sheet starting with '{sheet_prefix}' found — skipping.", file=sys.stderr)
            continue

        ws = wb[sheet_name]
        print(f"\n🔬 Processing sheet '{sheet_name}' → school={school_code}, has_afternoon={has_afternoon}")
        # Use a copy of the worksheet for expansion (avoid mutating original)
        # Actually openpyxl's cell.value setter works on the live worksheet — that's fine for our read-only purpose.
        sheet_entries = process_sheet(ws, school_code, has_afternoon)
        print(f"   Parsed {len(sheet_entries)} entries from '{sheet_name}'")
        all_entries.extend(sheet_entries)

    print(f"\n📊 Total entries before sort: {len(all_entries)}")

    # Sort
    all_entries = sort_entries(all_entries)

    # Validate
    print("\n🔍 Running validation checks...")
    issues = validate_entries(all_entries)
    errors = [i for i in issues if not i.startswith("Duplicate")]
    warnings = [i for i in issues if i.startswith("Duplicate") or "courseCode" in i]

    if errors:
        print("\n❌ VALIDATION ERRORS (will NOT write output):")
        for e in errors:
            print(f"   {e}")
        return 2
    if warnings:
        print("\n⚠️  VALIDATION WARNINGS (will write output anyway):")
        for w in warnings:
            print(f"   {w}")
    else:
        print("   ✅ All checks passed.")

    # Print summary
    print(f"\n✅ Parsed {len(all_entries)} entries:")
    by_school = {}
    by_date = {}
    for e in all_entries:
        by_school[e["school"]] = by_school.get(e["school"], 0) + 1
        by_date[e["date"]] = by_date.get(e["date"], 0) + 1
    print(f"   By school: {dict(sorted(by_school.items()))}")
    print(f"   By date:   {dict(sorted(by_date.items()))}")
    print(f"   First:    {all_entries[0]['date']} {all_entries[0]['day']} — {all_entries[0]['courseCode']} {all_entries[0]['courseName']}")
    print(f"   Last:     {all_entries[-1]['date']} {all_entries[-1]['day']} — {all_entries[-1]['courseCode']} {all_entries[-1]['courseName']}")

    if args.dry_run:
        print("\n🌵 Dry-run mode: output NOT written.")
        return 0

    # Write output (with safety guard: never overwrite with empty)
    if not all_entries:
        print("\n❌ Refusing to write empty array (safety guard).", file=sys.stderr)
        return 3

    output_path.parent.mkdir(parents=True, exist_ok=True)
    # Make a backup of existing file if present
    if output_path.exists():
        backup_path = output_path.with_suffix(".json.bak")
        backup_path.write_text(output_path.read_text())
        print(f"\n💾 Backed up existing file to: {backup_path}")

    output_path.write_text(json.dumps(all_entries, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(f"\n✅ Wrote {len(all_entries)} entries to: {output_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
